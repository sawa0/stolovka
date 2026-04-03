from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, json

def create_excel_report(data):
    
    if data == []:
        return
    
    formated = {}

    for order in data:
        order = list(order)
        order[5] = json.loads(order[5])
        
        if order[4] in formated:
            formated[order[4]][1].append(order[:3] + order[5:])
        else:
            formated[order[4]] = [order[3], [order[:3] + order[5:]]]

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Устанавливаем стиль для заголовков
    header_font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row = 1
    for key, value in formated.items():
        # Оставляем две пустые колонки перед клиентом
        user_column_offset = 1

        # Печать имени клиента и итога за месяц
        ws.cell(row=row, column=user_column_offset, value=value[0]).font = header_font
        ws.cell(row=row, column=user_column_offset + 1, value="Итог за месяц:").font = header_font
        month_total = sum(float(order[4]) for order in value[1])
        ws.cell(row=row, column=user_column_offset + 2, value=month_total).font = header_font

        # Индекс начала строк с подробностями для группировки
        start_detail_row = row + 1
        row += 1

        # Заголовки колонок для деталей
        headers = ["Order ID", "Дата", "Время", "Блюдо", "Цена", "Количество", "Итого"]
        for col_num, header in enumerate(headers, user_column_offset):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = alignment
            cell.border = thin_border
        row += 1

        for order in value[1]:
            order_id, date, time, dishes, order_total = order
            for dish in dishes:
                dish_name, (price, quantity) = dish
                ws.cell(row=row, column=user_column_offset, value=order_id)
                ws.cell(row=row, column=user_column_offset + 1, value=date)
                ws.cell(row=row, column=user_column_offset + 2, value=time)
                ws.cell(row=row, column=user_column_offset + 3, value=dish_name)
                ws.cell(row=row, column=user_column_offset + 4, value=float(price))
                ws.cell(row=row, column=user_column_offset + 5, value=int(quantity))
                ws.cell(row=row, column=user_column_offset + 6, value=float(price) * int(quantity))
                row += 1

        # Группировка строк с подробностями
        ws.row_dimensions.group(start_detail_row, row - 1, hidden=True)
        row += 1  # Дополнительный отступ между пользователями

    # Устанавливаем авторазмер колонок
    for col in range(user_column_offset, user_column_offset + 7):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    # Включаем сворачивание групп и устанавливаем свойство для сворачивания
    ws.sheet_properties.outlineSummaryBelow = False

    # Формируем название файла с учётом даты и времени
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"report_{timestamp}.xlsx"

    # Создаём папку reports, если она не существует
    reports_dir = "reports"
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)

    # Полный путь для сохранения файла в папку reports
    file_path = os.path.join(reports_dir, filename)

    # Проверка на существование файла и сохранение
    while os.path.exists(file_path):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(reports_dir, f"report_{timestamp}.xlsx")

    wb.save(file_path)
    
    return filename

def create_excel_recipe(data):
    if not data:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Рецепт"

    # Стили
    header_font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # Название блюда
    dish_name = data.get('Name', [''])[0]
    ws.cell(row=row, column=1, value=f"Блюдо: {dish_name}").font = header_font
    row += 2

    # Заголовки
    headers = ["Ингредиент", "Норма", "Цена", "Сумма"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = alignment
        cell.border = thin_border
    row += 1

    recipe = data.get('Recipe', {})
    ingredients = data.get('ingridients', {})

    total_cost = 0
    total_kg = 0
    total_l = 0

    start_detail_row = row

    for ing_id_str, amount in recipe.items():
        ing_id = int(ing_id_str)

        ing_data = ingredients.get(ing_id)
        if not ing_data:
            continue

        name = ing_data.get('name', '')
        volume = ing_data.get('volume', '')
        price = float(ing_data.get('price', 0))

        amount = float(amount)
        norm = f"{amount} {volume}"

        cost = price * amount
        total_cost += cost

        if volume == 'кг.':
            total_kg += amount
        elif volume == 'л.':
            total_l += amount

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=norm)
        ws.cell(row=row, column=3, value=price)
        ws.cell(row=row, column=4, value=cost)

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border

        row += 1

    # 👉 Общий объём (умный вывод)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value="Общий объём ингредиентов")
    cell.font = header_font
    cell.alignment = Alignment(horizontal="right")

    value_row = row

    if total_kg > 0:
        ws.cell(row=value_row, column=4, value=f"{round(total_kg, 3)} кг").font = header_font
        value_row += 1

    if total_l > 0:
        ws.cell(row=value_row, column=4, value=f"{round(total_l, 3)} л").font = header_font
        value_row += 1

    # если вдруг оба 0 (маловероятно, но на всякий)
    if total_kg == 0 and total_l == 0:
        ws.cell(row=value_row, column=4, value="0").font = header_font
        value_row += 1

    row = value_row

    # 👉 Себестоимость (уже с учётом сдвига)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value="Себестоимость")
    cell.font = header_font
    cell.alignment = Alignment(horizontal="right")

    ws.cell(row=row, column=4, value=round(total_cost, 2)).font = header_font

    # Группировка
    ws.row_dimensions.group(start_detail_row, row - 2, hidden=False)

    # Авторазмер
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    ws.sheet_properties.outlineSummaryBelow = False

    # Имя файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"recipe_{timestamp}.xlsx"

    reports_dir = "reports"
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)

    file_path = os.path.join(reports_dir, filename)

    while os.path.exists(file_path):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(reports_dir, f"recipe_{timestamp}.xlsx")

    wb.save(file_path)

    return filename
